from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText #生成指定格式的内容
from email.header import Header

#加载excel
wb = load_workbook('./21批量发送工资条/aa.xlsx')
#获取当前激活的工作簿
sheet1 = wb.active

# #获取行数据
count = 0
#记录表头
th = ''
for row in sheet1.iter_rows():
    # for c in row:
    #     print(c.value)
    if count == 0:
        th = f'''
        <tr>
            <td>{row[0].value}</td>
            <td>{row[1].value}</td>
            <td>{row[2].value}</td>
            <td>{row[3].value}</td>
            <td>{row[4].value}</td>
            <td>{row[5].value}</td>
            <td>{row[6].value}</td>
            <td>{row[7].value}</td>
            <td>{row[8].value}</td>
            <td>{row[9].value}</td>
        </tr>'''
        count += 1
        continue
    #msg = f'id:{row[0].value} name:{row[1].value} 部门:{row[2].value} 基本工资:{row[3].value} 应发工资:{row[9].value}'
    msg = f'''
        <tr>
            <td>{row[0].value}</td>
            <td>{row[1].value}</td>
            <td>{row[2].value}</td>
            <td>{row[3].value}</td>
            <td>{row[4].value}</td>
            <td>{row[5].value}</td>
            <td>{row[6].value}</td>
            <td>{row[7].value}</td>
            <td>{row[8].value}</td>
            <td>{row[9].value}</td>
        </tr>'''
    #发送邮件
    #邮件的地址服务器是哪个
    smtp_obj = smtplib.SMTP('smtp.qq.com')
    #登录邮箱,授权密码
    smtp_obj.login('1563562058@qq.com','xjskuxwuvjovijfa')
    #定义要发送的内容
    msg_body = MIMEText('<table border="1"'+th+msg+'</table>','html','utf-8')
    #定义从哪里发来的邮件
    msg_body['From'] = Header('测试部门发送来的邮件','utf-8')
    #邮件的主题
    msg_body['Subject'] = Header('第一次发送邮件','utf-8')
    #发送邮件
    smtp_obj.sendmail('1563562058@qq.com',[row[10].value],msg_body.as_string())
    print(f'已成功发送给{row[1].value}')